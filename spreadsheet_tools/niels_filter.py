import openpyxl
import csv
from datetime import datetime
import argparse


DEBUG = True
if DEBUG: print 'DEBUGGING MODE'

class PowerString(object):
    def __init__(self, start_date, power_string):
        self.start_date = start_date
        self.power_string = power_string

    def add_reading(self, reading):
        self.power_string.append(float(reading))

    def count(self):
        return len(self.power_string)

    def hours(self):
        return ((self.count() * 5.0) / 60.0)

    def total_kwh(self):
        return sum(self.power_string)

    def total_kwhph(self):
        return self.total_kwh() / self.hours()

    def __repr__(self):
        return "Count: {}, Hours: {}, Total kW-H: {}".format(self.count(), self.hours(), self.total_kwh())


class PowerStringAnalyzer(object):
    def __init__(self, power_strings):
        self.power_strings = power_strings

    def avg_kwhph(self):
        vals = []
        for e in self.power_strings:
            vals.append(e.total_kwhph())
        return average(vals)

    def avg_kwh(self):
        vals = []
        for e in self.power_strings:
            vals.append(e.total_kwh())
        return average(vals)

    def avg_hours(self):
        vals = []
        for e in self.power_strings:
            vals.append(e.hours())
        return average(vals)

    def avg_count(self):
        vals = []
        for e in self.power_strings:
            vals.append(e.count())
        return average(vals)

    def avg_l(self):
        return [self.avg_kwhph(), self.avg_kwh(), self.avg_hours(), self.avg_count()]


class PowerReading(object):
    def __init__(self, reading, date):
        self.reading = reading
        self.date = date

    def __repr__(self):
        return "{} {}".format(self.reading, self.date)


def open_power_file(file_name, header_l):
    csvfile = open(file_name)
    # with open(file_name) as csvfile:
    #     # Skips the none-header lines
    for line in xrange(header_l - 1):
        next(csvfile)

    dict_reader = csv.DictReader(csvfile) # Dict reader with proper header loaded (hopefully)

    return dict_reader


def extract_power_strings(dict_reader, col_name, min_pow, min_instncs):
    power_strings = []
    current_string = []
    chain_count = 0

    for row in dict_reader:
        current_date_time = "{} {}".format(row['Date'], row['End Time'])
        current_pow = float(row[col_name])

        if current_pow > 10: # Making sure ridiculous values don't get through
            continue

        power_reading = PowerReading(reading=current_pow, date=current_date_time)

        # print chain_count

        # Triggers if a power chain has ended.
        # Let's see if the chain is substantial enough to be made into a PowerString object.
        if current_pow < min_pow:

            # Too short. Drop the PowerString in progress
            if chain_count < min_instncs:
                chain_count = 0
                current_string = []

            # Satisfactory length. Create new PowerString object
            elif chain_count >= min_instncs:
                new_power_string = PowerString(
                    power_string=power_readings_to_list(current_string),
                    start_date=current_string[0].date
                )
                power_strings.append(new_power_string)

                # We still need to reset the counter.
                chain_count = 0
                current_string = []

        # Triggers every time we encounter a satisfactory power value.
        elif current_pow >= min_pow:
            chain_count += 1
            current_string.append(power_reading)

    # For - else loop. Else triggers once file is completely read.
    else:
        if chain_count >= min_instncs:
            new_power_string = PowerString(
                power_string=power_readings_to_list(current_string),
                start_date=current_string[0].date
            )
            power_strings.append(new_power_string)

    if DEBUG:
        for e in power_strings: print(e)

    return power_strings


def powerstring_to_excel(power_strings, file_name, min_pow, min_instncs):
    # Open and define our Excel sheets
    wb = openpyxl.Workbook()
    analysis = wb.active
    analysis.title = "Analysis"
    overview = wb.create_sheet("Overview")

    out_book = remove_extensions(file_name) + "_ANALYSIS.XLSX"

    write_analysis_header(analysis, min_pow, min_instncs)
    write_energy_data(analysis, power_strings)
    write_overview_data(overview, power_strings)

    wb.save(out_book)

    if DEBUG:
        print analysis['A1'].value
        for sheet in wb: print sheet


def write_energy_data(sheet, power_strings):
    current_column = 2
    current_row = 5
    num_data_fields = 5 # Refers to the number of headings applied by the write_analysis_header earlier
    start_row = current_row + num_data_fields
    num_datasets = len(power_strings)

    # Begin writing out data to sheet
    for dataset in power_strings:
        write_powerstring_header(sheet, dataset, current_column, current_row)
        # Going from the end of the powerstring headers to the final row of the dataset
        for row in xrange(start_row, start_row + dataset.count()):
            sheet.cell(column=current_column, row=row).value = dataset.power_string[row - start_row]

        current_column += 1 # Write next data to the next column to the right


def write_overview_data(sheet, power_strings):
    analyzer = PowerStringAnalyzer(power_strings)
    headers = ['AVG kW-Hr/hr', 'AVG kW-hr', 'AVG hours', 'AVG count']

    for index, e in enumerate(analyzer.avg_l()):
        sheet.cell(column=1, row=index + 1).value = headers[index]
        sheet.cell(column=2, row=index + 1).value = e


def write_powerstring_header(sheet, power_string, col, row):
    sheet.cell(column=col, row=row).value = power_string.total_kwhph()
    sheet.cell(column=col, row=row + 1).value = power_string.total_kwh()
    sheet.cell(column=col, row=row + 2).value = power_string.hours()
    sheet.cell(column=col, row=row + 3).value = power_string.count()
    sheet.cell(column=col, row=row + 4).value = power_string.start_date


def write_analysis_header(sheet, min_pow, min_instncs):
    # Legend
    sheet['A1'] = 'Date-Time: Date/Time info marking beginning of a significant ({} >{}kW-H energy instances chained together) energy usage period'.format(min_instncs, min_pow)
    sheet['A2'] = 'Hours: Length of time for significant energy usage period'
    sheet['A3'] = 'Count: Number of instances of significant energy usages'

    # Tabular info
    sheet['A5'] = 'kW-Hr/hr'
    sheet['A6'] = 'Total kW-hr'
    sheet['A7'] = 'hours'
    sheet['A8'] = 'count'
    sheet['A9'] = 'Date-Time'
    sheet['A10'] = 'kW-Hr / 5 minutes'


def power_readings_to_list(power_readings):
    power_string = []
    for e in power_readings:
        power_string.append(e.reading)

    return power_string


def remove_extensions(s):
    period = s.find('.')
    return s[:period]


def average(l):
    return sum(l) / len(l)



if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Power file analyzer.')
    parser.add_argument('-file', type=str, help="Name of the CSV file.")
    parser.add_argument('-header', type=int, help="Line which contains the header.")
    parser.add_argument('-col', type=str, help="Column to check for power values.")
    parser.add_argument('-min_pow', type=float, help="Values below this will not be considered for analysis.")
    parser.add_argument('-min_instncs', type=int, help="Minimum number of power usage instances for a column to be kept")

    args = parser.parse_args()

    now = datetime.now()
    dict_reader = open_power_file(args.file, args.header)
    power_strings = extract_power_strings(dict_reader, args.col, args.min_pow, args.min_instncs)
    excel = powerstring_to_excel(power_strings, args.file, args.min_pow, args.min_instncs)

    runtime = datetime.now() - now
    print runtime



